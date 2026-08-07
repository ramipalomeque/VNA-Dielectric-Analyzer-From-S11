from datetime import datetime
from openpyxl import load_workbook
import numpy as np

import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter

##
# @brief Valida y normaliza una fecha de medición.
#
# Convierte una fecha ingresada en distintos formatos a un formato único
# `dd-MM-yy`, utilizado para localizar las carpetas de mediciones y resultados.
#
# Formatos aceptados:
# - dd-MM-yy
# - dd/MM/yy
# - dd-MM-yyyyclar
# - dd/MM/yyyy
# - yyyy-MM-dd
# - yyyy/MM/dd
# - ddMMyy
# - ddMMyyyy
#
# @param date_str Fecha ingresada por el usuario.
#
# @return Fecha normalizada en formato `dd-MM-yy`.
#
# @exception ValueError Si la fecha no corresponde a ninguno de los formatos
#            soportados.
##
def parse_measurement_date(date_str):
    formatos = [
        "%d-%m-%y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d%m%y",
        "%d%m%Y",
    ]

    for formato in formatos:
        try:
            fecha = datetime.strptime(date_str.strip(), formato)
            return fecha.strftime("%d-%m-%y")
        except ValueError:
            pass

    raise ValueError(
        "Formato de fecha inválido. Ejemplos válidos: "
        "24-11-26, 24/11/2026, 2026-11-24."
    )


##
# @brief Exporta datos complejos utilizando una hoja template de Excel.
#
# Copia la hoja base "Hoja1" del template, la renombra con el nombre indicado
# y completa los datos de frecuencia, parte real y parte imaginaria.
#
# @param frecs Vector de frecuencias.
# @param complex_values Vector de valores complejos.
# @param sheet_name Nombre de la hoja destino.
# @param output_path Archivo Excel resultado.
# @param wb Objeto Excel.
#
# @return None.
##
def export_complex_to_excel(frecs, complex_values, sheet_name, wb):

    if "Hoja1" not in wb.sheetnames:
        raise Exception("No existe la hoja template Hoja1")

    ws_template = wb["Hoja1"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.copy_worksheet(ws_template)
    ws.title = sheet_name

    for index, (freq, value) in enumerate(zip(frecs, complex_values), start=2):
        ws.cell(index, 1).value = freq
        ws.cell(index, 2).value = np.real(value)
        ws.cell(index, 3).value = np.imag(value)

def open_workbook(file_path):
    return load_workbook(file_path)

def save_workbook(file_path, wb):
    ws = wb["Hoja1"]
    wb.remove(ws)
    wb.save(file_path)

def close_workbook(wb):
    wb.close()
    
def plot_er(frecs, er, title, output_file):
    frecs_MHz = frecs / 1e6

    plt.figure()
    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    plt.plot(frecs_MHz, np.real(er), label='Parte Real')
    plt.plot(frecs_MHz, np.imag(er), label='Parte Imaginaria')

    plt.title(title)
    plt.xlabel('Frecuencia (MHz)')
    plt.ylabel('Er')

    ax = plt.gca()
    plt.xscale('log')
    ax.xaxis.set_major_formatter(ScalarFormatter())

    plt.legend()
    plt.grid(True, which='major', linewidth=1)
    plt.grid(True, which='minor', linewidth=0.4)

    plt.savefig(output_file, dpi=300, bbox_inches='tight')

def plot_s11(frecs, s11, title, output_file):
    
    frecs_MHz = frecs / 1e6

    plt.figure()
    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    plt.plot(frecs_MHz, np.real(s11), label='Parte Real')
    plt.plot(frecs_MHz, np.imag(s11), label='Parte Imaginaria')

    plt.title(title)
    plt.xlabel('Frecuencia (MHz)')
    plt.ylabel('S11')

    ax = plt.gca()
    plt.xscale('log')
    ax.xaxis.set_major_formatter(ScalarFormatter())

    plt.legend()
    plt.grid(True, which='major', linewidth=1)
    plt.grid(True, which='minor', linewidth=0.4)

    plt.savefig(output_file, dpi=300, bbox_inches='tight')


def plot_s11_db_phase(frecs, s11, title, output_file=None):

    frecs_MHz = frecs / 1e6
    modulo_db = 20 * np.log10(np.abs(s11))
    fase_deg = np.angle(s11, deg=True)

    fig, (ax_mod, ax_phase) = plt.subplots(2, 1, figsize=(16, 9))

    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    fig.suptitle(title)
    ax_mod.plot(frecs_MHz, modulo_db)
    ax_mod.set_title("Módulo")
    #ax_mod.set_xlabel("Frecuencia (MHz)")
    ax_mod.set_ylabel("|S11| (dB)")
    ax_mod.set_xscale("log")
    ax_mod.xaxis.set_major_formatter(ScalarFormatter())
    ax_mod.grid(True, which="major", linewidth=1)
    ax_mod.grid(True, which="minor", linewidth=0.4)

    ax_phase.plot(frecs_MHz, fase_deg)
    ax_phase.set_title("Fase")
    ax_phase.set_xlabel("Frecuencia (MHz)")
    ax_phase.set_ylabel("Fase (°)")
    ax_phase.set_xscale("log")
    ax_phase.xaxis.set_major_formatter(ScalarFormatter())
    ax_phase.grid(True, which="major", linewidth=1)
    ax_phase.grid(True, which="minor", linewidth=0.4)

    fig.tight_layout()

    if output_file is not None:
        plt.savefig(output_file, dpi=300, bbox_inches='tight')
##
# @brief Grafica el parámetro S11 de una lista de mediciones.
#
# Genera un gráfico comparativo del parámetro S11 en función de la frecuencia
# para múltiples mediciones. Permite visualizar diferentes respuestas de la
# sonda o materiales bajo prueba en un mismo gráfico.
#
# Cada medición se representa utilizando el módulo y la fase del parámetro S11.
# El gráfico utiliza escala logarítmica en frecuencia y permite guardar la
# imagen generada en un archivo.
#
# @param frecs Vector de frecuencias en Hz.
# @param s11s Lista de diccionarios con las mediciones S11.
#        Cada elemento debe contener:
#        - name: Nombre de la medición.
#        - s11: Vector complejo correspondiente al parámetro S11.
# @param title Título del gráfico.
# @param output_file Ruta del archivo donde se guardará la imagen generada.
#        Si es None, solamente se muestra el gráfico.
#
# @return None.
##
def plot_er_list(frecs, ers, title, output_file = None):
    frecs_MHz = frecs / 1e6

    fig, (ax_real, ax_imag) = plt.subplots(2, 1, figsize=(16, 9), sharex=True, constrained_layout=True)
    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    for er in ers:
        line, = ax_real.plot(frecs_MHz, np.real(er["er"]), label=er["name"])
        ax_imag.plot(frecs_MHz, np.imag(er["er"]), color=line.get_color(), label=er["name"])


    fig.suptitle(title)
    ax_real.set_title("Parte Real")
    ax_real.set_ylabel("Re{er}")  
    ax_real.set_xscale("log")
    ax_real.xaxis.set_major_formatter(ScalarFormatter())
    ax_real.tick_params(axis='x', which='both', labelbottom=True)
    ax_real.grid(True, which="major", linewidth=1)
    ax_real.grid(True, which="minor", linewidth=0.4)
    ax_real.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    ax_imag.set_title("Parte Imaginaria")
    ax_imag.set_xlabel("Frecuencia (MHz)")
    ax_imag.set_ylabel("Im{er}")
    ax_imag.set_xscale("log")
    ax_imag.xaxis.set_major_formatter(ScalarFormatter())
    ax_imag.grid(True, which="major", linewidth=1)
    ax_imag.grid(True, which="minor", linewidth=0.4)
    ax_imag.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    if output_file is not None:
        plt.savefig(output_file, dpi=300, bbox_inches='tight') 

##
# @brief Grafica el parámetro S11 de una lista de mediciones.
#
# Genera un gráfico comparativo del parámetro S11 en función de la frecuencia
# para múltiples mediciones. Permite visualizar diferentes respuestas de la
# sonda o materiales bajo prueba en un mismo gráfico.
#
# Cada medición se representa utilizando el módulo y la fase del parámetro S11.
# El gráfico utiliza escala logarítmica en frecuencia y permite guardar la
# imagen generada en un archivo.
#
# @param frecs Vector de frecuencias en Hz.
# @param s11s Lista de diccionarios con las mediciones S11.
#        Cada elemento debe contener:
#        - name: Nombre de la medición.
#        - s11: Vector complejo correspondiente al parámetro S11.
# @param title Título del gráfico.
# @param output_file Ruta del archivo donde se guardará la imagen generada.
#        Si es None, solamente se muestra el gráfico.
#
# @return None.
##
def plot_s11_list(frecs, s11s, title, output_file=None):

    frecs_MHz = frecs / 1e6

    fig, (ax_real, ax_imag) = plt.subplots(2, 1, figsize=(16, 9), sharex=True, constrained_layout=True)

    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    for s11 in s11s:
        line, = ax_real.plot(frecs_MHz, np.real(s11["s11"]), label=s11["name"])
        ax_imag.plot(frecs_MHz, np.imag(s11["s11"]), color=line.get_color(), label=s11["name"])

    fig.suptitle(title)
    ax_real.set_title("Parte Real")
    ax_real.set_ylabel("Re{S11}")  
    ax_real.set_xscale("log")
    ax_real.xaxis.set_major_formatter(ScalarFormatter())
    ax_real.tick_params(axis='x', which='both', labelbottom=True)
    ax_real.grid(True, which="major", linewidth=1)
    ax_real.grid(True, which="minor", linewidth=0.4)
    ax_real.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    ax_imag.set_title("Parte Imaginaria")
    ax_imag.set_xlabel("Frecuencia (MHz)")
    ax_imag.set_ylabel("Im{S11}")
    ax_imag.set_xscale("log")
    ax_imag.xaxis.set_major_formatter(ScalarFormatter())
    ax_imag.grid(True, which="major", linewidth=1)
    ax_imag.grid(True, which="minor", linewidth=0.4)
    #ax_imag.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    if output_file is not None:
        plt.savefig(output_file, dpi=300, bbox_inches='tight')


def plot_s11_db_phase_list(frecs, s11s, title, output_file=None):

    frecs_MHz = frecs / 1e6

    fig, (ax_mod, ax_phase) = plt.subplots(2, 1, figsize=(16, 9), sharex=True, constrained_layout=True)

    manager = plt.get_current_fig_manager()
    manager.window.state('zoomed')

    fig.suptitle(title)

    for s11 in s11s:
        modulo_db = 20 * np.log10(np.abs(s11["s11"]))
        fase_deg = np.angle(s11["s11"], deg=True)

        line, = ax_mod.plot(frecs_MHz, modulo_db, label=s11["name"])
        ax_phase.plot(frecs_MHz, fase_deg, color=line.get_color(), label=s11["name"])
        

    ax_mod.set_title("Módulo")
    ax_mod.set_xlabel("Frecuencia (MHz)")
    ax_mod.set_ylabel("|S11| (dB)")
    ax_mod.set_xscale("log")
    ax_mod.xaxis.set_major_formatter(ScalarFormatter())
    ax_mod.grid(True, which="major", linewidth=1)
    ax_mod.grid(True, which="minor", linewidth=0.4)
    ax_mod.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    ax_phase.set_title("Fase")
    ax_phase.set_xlabel("Frecuencia (MHz)")
    ax_phase.set_ylabel("Fase (°)")
    ax_phase.set_xscale("log")
    ax_phase.xaxis.set_major_formatter(ScalarFormatter())
    ax_phase.grid(True, which="major", linewidth=1)
    ax_phase.grid(True, which="minor", linewidth=0.4)
    ax_phase.legend(loc="upper left", bbox_to_anchor=(1.02, 1), ncol=1, fontsize=12)

    if output_file is not None:
        plt.savefig(output_file, dpi=300, bbox_inches="tight")
def show_plots():
    plt.show()